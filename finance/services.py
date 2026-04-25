"""
finance/services.py
Logique métier : nomenclatures automatiques, calculs TVA, cotisations URSSAF.
"""
from decimal import Decimal, ROUND_HALF_UP
from datetime import date

from config_app.models import ParametreCotisation


# ─── Numérotation automatique ────────────────────────────────────────────────

def generate_numero_facture_vente(type_facture, annee: int, mois: int, suffixe: str = '') -> str:
    """
    Génère le numéro de facture de vente.
    Nomenclature : FV + AA + MM + NN + _[Type]
    Subvention : S + AA + MM + NN
    """
    from .models import FactureVente

    is_sub = type_facture.est_subvention
    prefix = 'S' if is_sub else 'FV'
    aa = str(annee)[-2:]
    mm = f"{mois:02d}"
    
    # On cherche l'incrément le plus haut pour ce préfixe et ce mois
    base_prefix = f"{prefix}{aa}{mm}"
    existing = FactureVente.objects.filter(numero__startswith=base_prefix).values_list('numero', flat=True)
    
    max_nn = 0
    for num in existing:
        try:
            # Extraction des 2 chiffres après le préfixe temporel
            nn_str = num[len(base_prefix):len(base_prefix)+2]
            nn = int(nn_str)
            if nn > max_nn: max_nn = nn
        except: continue
    
    chrono = max_nn + 1
    nn = f"{chrono:02d}"
    base_ref = f"{prefix}{aa}{mm}{nn}"
    
    if is_sub:
        return base_ref
        
    suffix_map = {'A': '_A', 'S': '_S', 'C': '_C', 'R': '_REF', 'AV': '_AV'}
    ext = suffix_map.get(suffixe, f"_{suffixe}" if suffixe else "")
    return f"{base_ref}{ext}"


def generate_numero_facture_achat(type_achat, annee: int, mois: int) -> str:
    """
    Génère le numéro de facture d'achat.
    Nomenclature : A (fournisseur) ou NF (ndf) + AA + MM + NN (partagé)
    """
    from .models import FactureAchat

    # Même logique : on cherche le chrono le plus élevé (tous préfixes confondus pour les achats)
    aa = str(annee)[-2:]
    mm = f"{mois:02d}"
    
    # On cherche dans les achats commençant par A ou NF pour ce mois
    # Note: comme le chrono est partagé, on regarde les deux préfixes possibles
    # On simplifie en cherchant par date d'opération si le numéro a été sali
    existing = FactureAchat.objects.filter(
        date_operation__year=annee,
        date_operation__month=mois
    ).values_list('numero', flat=True)
    
    max_nn = 0
    for num in existing:
        try:
            # On cherche les 2 chiffres après le préfixe temporel (index 5-7 pour A2404XX ou 6-8 pour NF2404XX)
            p_len = 5 if num.startswith('A') else 6
            nn_str = num[p_len:p_len+2]
            nn = int(nn_str)
            if nn > max_nn: max_nn = nn
        except: continue
        
    chrono = max_nn + 1
    nn = f"{chrono:02d}"
    prefix = 'NF' if type_achat.suffixe == 'NF' else 'A'
    return f"{prefix}{aa}{mm}{nn}"


def generate_numero_bv(annee: int) -> str:
    """
    Génère le numéro de BV.
    Format : BV_YY-NN (ex: BV_26-01), incrémental sur l'année.
    Unicité garantie.
    """
    from .models import BulletinVersement

    aa = str(annee)[-2:]
    prefix = f"BV_{aa}-"

    # Trouver le plus grand index existant pour ce préfixe
    existing = BulletinVersement.objects.filter(
        numero__startswith=prefix
    ).order_by('-numero')

    if existing.exists():
        last_num = existing.first().numero
        try:
            # Extraction du nombre après le tiret
            last_idx = int(last_num.split('-')[-1])
        except (ValueError, IndexError):
            last_idx = 0
        new_idx = last_idx + 1
    else:
        new_idx = 1

    return f"{prefix}{new_idx:02d}"


def to_decimal(value):
    """Convertit une chaîne (éventuellement avec virgule) en Decimal."""
    if not value:
        return Decimal('0')
    if isinstance(value, str):
        value = value.replace(',', '.')
    try:
        return Decimal(value)
    except:
        return Decimal('0')


# ─── Calculs TVA ─────────────────────────────────────────────────────────────

def calculate_tva(montant_ttc: Decimal, taux: Decimal) -> dict:
    """
    Calcule HT et TVA à partir du TTC et du taux.

    Returns:
        {'ht': Decimal, 'tva': Decimal}
    """
    taux_decimal = taux / Decimal('100')
    ht = (montant_ttc / (1 + taux_decimal)).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
    tva = (montant_ttc - ht).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
    return {'ht': ht, 'tva': tva}


def calculate_ht_tva(montant_ht: Decimal, taux: Decimal) -> dict:
    """
    Calcule TVA et TTC à partir du HT et du taux.
    """
    taux_decimal = taux / Decimal('100')
    tva = (montant_ht * taux_decimal).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
    ttc = (montant_ht + tva).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
    return {'tva': tva, 'ttc': ttc}


def get_taux_tva_defaut(type_facture) -> Decimal:
    """Retourne le taux de TVA par défaut selon le type de facture."""
    return Decimal(str(type_facture.taux_tva_defaut))


# ─── Calculs cotisations URSSAF ──────────────────────────────────────────────

def calculate_cotisations_urssaf(nb_jeh: Decimal) -> dict:
    """
    Calcule systématiquement les cotisations URSSAF pour la Part Junior (JE) 
    et la Part Étudiant (Intervenant).
    """
    from config_app.models import ParametreCotisation

    # Récupération des deux types de paramètres requis
    p_j = ParametreCotisation.objects.get(type_cotisant='junior')
    p_e = ParametreCotisation.objects.get(type_cotisant='etudiant')

    # L'assiette est basée sur la base URSSAF du profil junior
    assiette = (nb_jeh * p_j.base_urssaf).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)

    def calc(taux: Decimal) -> Decimal:
        return (assiette * taux / Decimal('100')).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)

    # --- Calcul Part Junior (JE) ---
    j_maladie = calc(p_j.assurance_maladie)
    j_at = calc(p_j.accident_travail)
    j_vp = calc(p_j.vieillesse_plafonnee)
    j_vd = calc(p_j.vieillesse_deplafonnee)
    j_af = calc(p_j.allocations_familiales)
    j_csgd = calc(p_j.csg_deductible)
    j_csgnd = calc(p_j.csg_non_deductible)
    total_j = j_maladie + j_at + j_vp + j_vd + j_af + j_csgd + j_csgnd

    # --- Calcul Part Étudiant (Intervenant) ---
    e_maladie = calc(p_e.assurance_maladie)
    e_at = calc(p_e.accident_travail)
    e_vp = calc(p_e.vieillesse_plafonnee)
    e_vd = calc(p_e.vieillesse_deplafonnee)
    e_af = calc(p_e.allocations_familiales)
    e_csgd = calc(p_e.csg_deductible)
    e_csgnd = calc(p_e.csg_non_deductible)
    total_e = e_maladie + e_at + e_vp + e_vd + e_af + e_csgd + e_csgnd

    return {
        'assiette': assiette,
        
        # Part Junior
        'j_maladie': j_maladie, 
        'j_at': j_at, 
        'j_vp': j_vp, 
        'j_vd': j_vd,
        'j_af': j_af, 
        'j_csgd': j_csgd, 
        'j_csgnd': j_csgnd, 
        'total_j': total_j,
        
        # Part Étudiant
        'e_maladie': e_maladie, 
        'e_at': e_at, 
        'e_vp': e_vp, 
        'e_vd': e_vd,
        'e_af': e_af, 
        'e_csgd': e_csgd, 
        'e_csgnd': e_csgnd, 
        'total_e': total_e,
        
        'total_global': total_j + total_e
    }


def generate_bv_pdf_from_template(data_dict: dict) -> bytes:
    """
    Remplit un template Excel avec des tags {{TAG}} et le convertit en PDF via LibreOffice.
    """
    import os
    import openpyxl
    import subprocess
    import tempfile
    import shutil
    from django.conf import settings

    template_path = os.path.join(settings.BASE_DIR, 'Ressource_gemini', 'doctype BV janvier 2026.xlsx')
    if not os.path.exists(template_path):
        raise FileNotFoundError(f"Modèle Excel introuvable à l'emplacement : {template_path}")

    # Création d'un répertoire temporaire pour le travail
    with tempfile.TemporaryDirectory() as tmp_dir:
        tmp_excel = os.path.join(tmp_dir, 'bulletin.xlsx')
        shutil.copy(template_path, tmp_excel)

        # Remplissage de l'Excel
        wb = openpyxl.load_workbook(tmp_excel)
        for sheet in wb.worksheets:
            for row in sheet.iter_rows():
                for cell in row:
                    if isinstance(cell.value, str):
                        for tag, value in data_dict.items():
                            if not isinstance(cell.value, str):
                                break
                            tag_str = f"{{{{{tag}}}}}"
                            if tag_str in cell.value:

                                # Remplacement du tag
                                if cell.value.strip() == tag_str:
                                    # On conserve le type (float/int) pour que les formules Excel fonctionnent
                                    cell.value = value
                                else:
                                    # Pour les remplacements partiels, on convertit en string
                                    cell.value = cell.value.replace(tag_str, str(value))

        # Réglages Impression / Zoom
        if 'Fiche BV' in wb.sheetnames:
            # On supprime tous les autres onglets pour ne garder que la Fiche BV
            for sn in wb.sheetnames:
                if sn != 'Fiche BV':
                    del wb[sn]
            
            ws_bv = wb['Fiche BV']
            
            # Supprimer ce qui dépasse de K60 pour être vraiment propre
            if ws_bv.max_column > 11:
                ws_bv.delete_cols(12, ws_bv.max_column - 11)
            if ws_bv.max_row > 60:
                ws_bv.delete_rows(61, ws_bv.max_row - 60)

            # Zone d'impression
            ws_bv.print_area = 'A1:K60'
            
            # Marges à zéro pour maximiser l'espace (Zoom)
            ws_bv.page_margins.left = 0
            ws_bv.page_margins.right = 0
            ws_bv.page_margins.top = 0
            ws_bv.page_margins.bottom = 0

            # Ajustement à la page
            ws_bv.page_setup.fitToPage = True
            ws_bv.page_setup.fitToHeight = 1
            ws_bv.page_setup.fitToWidth = 1
            ws_bv.page_setup.orientation = ws_bv.ORIENTATION_PORTRAIT
            ws_bv.page_setup.paperSize = ws_bv.PAPERSIZE_A4
        
        wb.save(tmp_excel)
        wb.close()



        # Conversion PDF via LibreOffice
        try:
            subprocess.run([
                'libreoffice', '--headless', '--convert-to', 'pdf',
                '--outdir', tmp_dir, tmp_excel
            ], check=True, capture_output=True, timeout=30)
            
            tmp_pdf = os.path.join(tmp_dir, 'bulletin.pdf')
            if not os.path.exists(tmp_pdf):
                raise Exception("Le fichier PDF n'a pas été généré (LibreOffice).")
            
            with open(tmp_pdf, 'rb') as f:
                return f.read()
                
        except subprocess.CalledProcessError as e:
            raise Exception(f"Erreur LibreOffice : {e.stderr.decode()}")
        except subprocess.TimeoutExpired:
            raise Exception("Le processus LibreOffice a expiré.")

def generate_ndf_pdf(ndf, sig_config) -> bytes:
    """Génère le PDF d'une NDF à partir d'un template Excel."""
    import os
    import openpyxl
    import subprocess
    import tempfile
    import shutil
    from django.conf import settings
    from datetime import date

    # Choix du template
    tpl_name = 'doctype_flower_NDF_kilometrique.xlsx' if ndf.type_frais == 'ik' else 'doctype_flower_NDF_achat.xlsx'
    template_path = os.path.join(settings.BASE_DIR, 'Ressource_gemini', tpl_name)
    
    if not os.path.exists(template_path):
        raise FileNotFoundError(f"Modèle Excel introuvable : {tpl_name}")

    # Préparation des données communes
    data = {
        'REF': ndf.facture_achat.numero if ndf.facture_achat else f"NDF-{ndf.id}",
        'DATE_GEN': date.today().strftime('%d/%m/%Y'),
        'NOM': ndf.nom_beneficiaire.upper(),
        'PRENOM': ndf.prenom_beneficiaire,
        'PRÉNOM': ndf.prenom_beneficiaire,
        'LIBELLE': ndf.libelle,
        'PRESIDENT': f"{sig_config.president_prenom} {sig_config.president_nom.upper()}",
        'TRESORIER': f"{sig_config.tresorier_prenom} {sig_config.tresorier_nom.upper()}",
        'JUSTIFICATIFS': ", ".join([doc.type_pièce or "Justificatif" for doc in ndf.justificatifs.all()]),
        'TOTAL_TTC': float(sum(l.montant_ttc for l in ndf.lignes.all())),
    }

    if ndf.type_frais == 'ik':
        ligne_ik = ndf.lignes.filter(est_ik=True).first()
        if ligne_ik:
            data['DISTANCE'] = float(ligne_ik.distance_km or 0)
            data['TOTAL_IK'] = float(ligne_ik.montant_ttc)
            # On suppose que le taux est TTC / Distance
            if ligne_ik.distance_km:
                data['PRIX_KM'] = float(ligne_ik.montant_ttc / ligne_ik.distance_km)
            else:
                data['PRIX_KM'] = 0.0
    else:
        # Achats : Totaux HT/TVA
        data['TOTAL_HT'] = float(sum(l.montant_ht for l in ndf.lignes.all()))
        data['TOTAL_TVA'] = float(sum(l.montant_tva for l in ndf.lignes.all()))
        # Lignes individuelles (jusqu'à 50)
        for i, ligne in enumerate(ndf.lignes.all()[:50], 1):
            data[f'DESC_{i}'] = ligne.libelle
            data[f'HT_{i}'] = float(ligne.montant_ht)
            data[f'TVA_{i}'] = float(ligne.montant_tva)
            data[f'TAUX_{i}'] = float(ligne.taux_tva)
            data[f'TTC_{i}'] = float(ligne.montant_ttc)

    # Travail temporaire
    with tempfile.TemporaryDirectory() as tmp_dir:
        tmp_excel = os.path.join(tmp_dir, 'ndf_temp.xlsx')
        shutil.copy(template_path, tmp_excel)
        wb = openpyxl.load_workbook(tmp_excel)
        
        for sheet in list(wb.worksheets):
            if sheet != wb.worksheets[0]:
                wb.remove(sheet)
                continue

            rows_to_delete = []
            for row in sheet.iter_rows():
                is_empty_desc_row = False
                for cell in row:
                    if isinstance(cell.value, str):
                        # 1. Remplacement des balises connues
                        match_found = False
                        for tag, value in data.items():
                            if not isinstance(cell.value, str):
                                break
                            tag_str = f"{{{{{tag}}}}}"
                            if tag_str in cell.value:
                                # Remplacement exact ou partiel
                                if cell.value.strip() == tag_str:
                                    cell.value = value
                                else:
                                    cell.value = cell.value.replace(tag_str, str(value))
                                match_found = True
                        
                        # Si c'est une balise de description non remplie, on marque la ligne
                        if not match_found and "{{DESC_" in cell.value:
                            is_empty_desc_row = True
                        
                        # 2. Nettoyage des balises restantes (non remplies)
                        if isinstance(cell.value, str) and "{{" in cell.value:
                            import re
                            cell.value = re.sub(r"\{\{.*?\}\}", "", cell.value)
                
                if is_empty_desc_row:
                    rows_to_delete.append(row[0].row)

            # Nouvelle méthode : Masquage total (Hidden + Height=0) 
            # On ne touche qu'aux lignes du tableau (19 à 68)
            from openpyxl.styles import Border, Side
            no_border = Border(left=Side(style=None), right=Side(style=None), top=Side(style=None), bottom=Side(style=None))
            
            for row_idx in range(1, 69): # On scanne la zone tableau
                is_empty = any(idx == row_idx for idx in rows_to_delete)
                if is_empty and row_idx >= 19: # On ne masque que si c'est une ligne DESC_n vide
                    sheet.row_dimensions[row_idx].hidden = True
                    sheet.row_dimensions[row_idx].height = 0
                    # On retire les bordures pour éviter les traits parasites en PDF
                    for cell in sheet[row_idx]:
                        cell.border = no_border

            # On laisse le bas de page (>= 69) totalement intact
            
            # Mise en page : on ajuste à la largeur
            sheet.page_setup.fitToPage = True
            sheet.page_setup.fitToWidth = 1
            sheet.page_setup.fitToHeight = 0
            sheet.page_setup.paperSize = sheet.PAPERSIZE_A4

        wb.save(tmp_excel)
        wb.close()

        # Conversion PDF de la note de frais
        subprocess.run(['libreoffice', '--headless', '--convert-to', 'pdf', '--outdir', tmp_dir, tmp_excel], 
                       check=True, capture_output=True, timeout=30)
        
        main_pdf = os.path.join(tmp_dir, 'ndf_temp.pdf')
        pdf_list = [main_pdf]
        
        # Traitement des justificatifs
        from PIL import Image
        for i, justif in enumerate(ndf.justificatifs.all()):
            if not justif.fichier or not os.path.exists(justif.fichier.path):
                continue
            
            ext = os.path.splitext(justif.fichier.path)[1].lower()
            if ext == '.pdf':
                pdf_list.append(justif.fichier.path)
            elif ext in ['.jpg', '.jpeg', '.png', '.bmp']:
                # Conversion image en PDF via PIL
                justif_pdf = os.path.join(tmp_dir, f'justif_{i}.pdf')
                try:
                    img = Image.open(justif.fichier.path)
                    if img.mode in ('RGBA', 'P'):
                        img = img.convert('RGB')
                    img.save(justif_pdf, "PDF", resolution=100.0)
                    pdf_list.append(justif_pdf)
                except Exception as e:
                    print(f"Erreur conversion image {justif.fichier.path}: {e}")

        if len(pdf_list) > 1:
            final_pdf = os.path.join(tmp_dir, 'final_ndf.pdf')
            try:
                # Utilisation de pdfunite (outil système) pour fusionner
                subprocess.run(['pdfunite'] + pdf_list + [final_pdf], check=True)
                tmp_pdf = final_pdf
            except Exception as e:
                print(f"Erreur fusion PDF: {e}")
                tmp_pdf = main_pdf
        else:
            tmp_pdf = main_pdf

        with open(tmp_pdf, 'rb') as f:
            return f.read()
def generate_numero_ndf(year, month):
    """Génère un numéro de type NF-YYMMXX avec reset mensuel."""
    from .models import FactureAchat
    
    prefix = f"NF-{str(year)[2:]}{month:02d}"
    
    # On cherche le dernier numéro commençant par ce prefixe
    last = FactureAchat.objects.filter(numero__startswith=prefix).order_by('-numero').first()
    
    if not last:
        return f"{prefix}01"
    
    try:
        # On extrait les 2 derniers chiffres
        last_num = int(last.numero[-2:])
        new_num = last_num + 1
        return f"{prefix}{new_num:02d}"
    except (ValueError, IndexError):
        return f"{prefix}01"
